import streamlit as st
import os
import subprocess
from datetime import datetime
import requests
import pytz
import base64
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np
import re




# URL of the API to get the current time
time_api_url = "http://worldtimeapi.org/api/timezone/Etc/UTC"

def get_current_time():
    try:
        response = requests.get(time_api_url)
        response.raise_for_status()
        current_time_data = response.json()
        return datetime.fromisoformat(current_time_data["datetime"]).replace(tzinfo=pytz.UTC)
    except requests.exceptions.RequestException as e:
        st.warning(f"Error fetching time: {e}. Using local system time.")
        return datetime.now(pytz.UTC)

def generate_ds_general(sn_value, df_mmsta):
    if pd.isna(sn_value):
        return ''
    start_rows = df_mmsta[df_mmsta['Material'] == sn_value]
    if start_rows.empty:
        return ''
    start_index = start_rows.index[0]
    ds_general_parts = []
    for i in range(start_index + 1, len(df_mmsta)):
        cell_value = df_mmsta.iloc[i]['Material']
        if pd.isna(cell_value):
            continue
        cell_value_str = str(cell_value)
        if cell_value_str.startswith('S00'):
            break
        if df_mmsta.iloc[i]['Material Description'].startswith('LT Single Wire'):
            component_quantity = df_mmsta.iloc[i]['Component quantity']
            ds_general_parts.append(f"{cell_value_str} ({component_quantity})")
        else:
            ds_general_parts.append(cell_value_str)
    return ' / '.join(ds_general_parts)

def separate_files(mmsta_file):
    df_mmsta = pd.read_excel(mmsta_file, sheet_name='MMSTA')
    
    def update_pn(row):
        level = extract_level(row['Level'])
        if level == 0:
            return row['Material Description']
        return None

    def extract_level(level_str):
        match = re.search(r'\d+', level_str)
        return int(match.group()) if match else None

    df_mmsta['PN'] = df_mmsta.apply(update_pn, axis=1)
    df_mmsta['PN'] = df_mmsta['PN'].fillna(method='ffill')

    def count_asterisks(value):
        return value.count('*')

    df_mmsta['asterisk_count'] = df_mmsta['Level'].apply(count_asterisks)
    max_asterisks = df_mmsta['asterisk_count'].max()
    n = max_asterisks + 1

    df_1 = df_mmsta[(df_mmsta['Level'] != 0) & (~df_mmsta['Level'].astype(str).str.endswith(str(max_asterisks))) & (df_mmsta['Material Type'] == 'YSFG')].copy()
    for i in range(1, n):
        df_1[f'SN{i}'] = np.nan
        df_1[f'DSN{i}'] = np.nan

    df_1.drop(columns=['asterisk_count'], inplace=True)

    def update_sn(row, level_indicator):
        if row['Level'] == level_indicator:
            return row['Material']
        return np.nan

    def update_and_fill_sn(df_1, sn_column, level_indicator):
        df_1.loc[:, sn_column] = df_1[sn_column].astype(object).ffill()  # Astuce pour éviter les conflits de types
        df_1[sn_column] = df_1[sn_column].fillna(method='ffill')

    def generate_level(n):
        return '*' * (n + 1) + str(n + 1)

    for i in range(n-1):
        column_name = f'SN{i + 1}'
        level = generate_level(i)
        update_and_fill_sn(df_1, column_name, level)

    material_description_map = df_1.set_index('Material')['Material Description'].to_dict()
    sn_columns = [f'SN{i}' for i in range(1, n)]
    dsn_columns = [f'DSN{i}' for i in range(1, n)]

    def fill_dsn_columns(df_1, sn_columns, dsn_columns, material_description_map):
        for sn_col, dsn_col in zip(sn_columns, dsn_columns):
            df_1.loc[:, dsn_col] = df_1[sn_col].map(material_description_map)

    fill_dsn_columns(df_1, sn_columns, dsn_columns, material_description_map)

    def clean_table(row):
        if row['Level'] == '*1':
            row['SN2'] = '(blank)'
            row['DSN2'] = '(blank)'
            row['SN3'] = '(blank)'
            row['DSN3'] = '(blank)'
        elif row['Level'] == '**2':
            row['SN3'] = '(blank)'
            row['DSN3'] = '(blank)'
        return row

    df_1 = df_1.apply(clean_table, axis=1)
    filtered_df = df_1.dropna(subset=['SN2', 'DSN2', 'SN3', 'DSN3'], how='all')
    columns_to_keep = ['SN1', 'DSN1', 'SN2', 'DSN2', 'SN3', 'DSN3', 'PN']
    filtered_df = filtered_df[columns_to_keep]

    grouped = filtered_df.groupby(['SN1', 'DSN1', 'SN2', 'DSN2', 'SN3', 'DSN3', 'PN']).size().reset_index(name='count')
    pivot_table = grouped.pivot(index=['SN1', 'DSN1', 'SN2', 'DSN2', 'SN3', 'DSN3'], columns='PN', values='count').fillna(0).reset_index()
    pivot_table.replace(0, np.nan, inplace=True)
    pivot_table['Total'] = pivot_table.sum(axis=1, numeric_only=True)

    def add_ds_general_column(sheet_name, df):
        if sheet_name == 'FIL SIMPLE':
            df.loc[:,'DS Général'] = df['SN1'].apply(lambda x: generate_ds_general(x, df_mmsta))
        elif sheet_name in ['joint', 'double', 'twist', 'SQUIB', 'GW']:
            df.loc[:,'DS Général'] = df['SN2'].apply(lambda x: generate_ds_general(x, df_mmsta))
        elif sheet_name == 'super group':
            df.loc[:,'DS Général'] = df['SN3'].apply(lambda x: generate_ds_general(x, df_mmsta))
        return df

    # Assurez-vous que 'DSN1' est de type chaîne de caractères avant d'appliquer .str.contains()
    pivot_table['DSN1'] = pivot_table['DSN1'].astype(str)
    
    filters = {
        'FIL SIMPLE': 'circuit ',
        'double': 'double',
        'twist': 'twisted',
        'joint': 'joint',
        'super group': 'super group',
        'SQUIB': 'simple super group',
        'cut tube': 'cut tube|GAFT',
        'GW': 'group wire'
    }


    with pd.ExcelWriter("MMSTA_separe.xlsx", engine='openpyxl') as writer:
        df_mmsta.to_excel(writer, sheet_name='MMSTA', index=False)
        pivot_table.to_excel(writer, sheet_name='SEPARER', index=False)

        for sheet_name, filter_str in filters.items():
            if sheet_name == 'super group':
                filtered_df = pivot_table[pivot_table['DSN1'].str.contains(filter_str, case=False, na=False) &
                                          ~pivot_table['DSN1'].str.contains('simple super group', case=False, na=False)]
            else:
                filtered_df = pivot_table[pivot_table['DSN1'].str.contains(filter_str, case=False, na=False)]

            filtered_df = add_ds_general_column(sheet_name, filtered_df)
            filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = openpyxl.load_workbook("MMSTA_separe.xlsx")
    sheet = workbook['SEPARER']
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in sheet[1]:
        cell.fill = yellow_fill
    workbook.save("MMSTA_separe.xlsx")

    return "MMSTA_separe.xlsx"

def integrate_files(separated_file, circuit_file):
    df_mmsta = pd.read_excel(separated_file, sheet_name='FIL SIMPLE')
    df_mmsta1 = pd.read_excel(separated_file, sheet_name='double')
    df_mmsta2 = pd.read_excel(separated_file, sheet_name='twist')
    df_mmsta3 = pd.read_excel(separated_file, sheet_name='SQUIB')
    df_mmsta4 = pd.read_excel(separated_file, sheet_name='GW')
    df_mmsta5 = pd.read_excel(separated_file, sheet_name='joint')
    df_mmsta6 = pd.read_excel(separated_file, sheet_name='super group')
    df_maxwire = pd.read_excel(circuit_file, sheet_name='Report(Draft)')

    insert_position = df_maxwire.columns.get_loc('Wire Internal Name') + 1
    df_maxwire.insert(insert_position, 'TYPE', '')
    df_maxwire.insert(insert_position + 1, 'SN ADD', '')
    df_maxwire.insert(insert_position + 2, 'SN P2', '')
    df_maxwire.insert(insert_position + 3, 'Super group', '')

    df_maxwire['Wire Internal Name'] = df_maxwire['Wire Internal Name'].astype(str).str.replace('W', '')
    df_maxwire['Wire Cross-Section'] = df_maxwire['Wire Cross-Section'].apply(lambda x: str(x).replace('.0', '') if str(x).endswith('.0') else str(x))
    df_maxwire['From Seal by Terminal'] = df_maxwire['From Seal by Terminal'].astype(str).str.replace('.0', '')
    df_maxwire['To Seal by Terminal'] = df_maxwire['To Seal by Terminal'].astype(str).str.replace('.0', '')
    df_maxwire['Wire Part Number'] = df_maxwire['Wire Part Number'].astype(str).str.replace('180', '')

    df_maxwire['salma'] = (
        'Circuit ' +
        df_maxwire['Wire Internal Name'].astype(str) + ' ' +
        df_maxwire['Wire Kind'].astype(str) + ' ' +
        df_maxwire['Wire Cross-Section'].astype(str) + ' ' +
        df_maxwire['Wire Color'].astype(str)
    )

    df_mmsta5_clean = df_mmsta5.dropna(subset=['SN2']).copy()
    df_mmsta5_clean = df_mmsta5_clean[df_mmsta5_clean['SN2'].str.strip() != '(blanks)']

    def map_sn_add_and_sn_p2(salma, df):
        matched_row = df[df['DSN2'] == salma]
        if not matched_row.empty:
            sn_add = matched_row['SN2'].tolist()
            sn_p2 = matched_row['SN1'].tolist()
            return sn_add, sn_p2
        return None, None

    def update_sn_add_and_sn_p2(row, df_mmsta5_clean):
        salma = row['salma']
        new_sn_add, new_sn_p2 = map_sn_add_and_sn_p2(salma, df_mmsta5_clean)

        if isinstance(row['SN ADD'], list) and any(str(val).strip() for val in row['SN ADD']):
            sn_add_value = row['SN ADD']
        else:
            sn_add_value = new_sn_add if new_sn_add else row['SN ADD']

        if isinstance(row['SN P2'], list) and any(str(val).strip() for val in row['SN P2']):
            sn_p2_value = row['SN P2']
        else:
            sn_p2_value = new_sn_p2 if new_sn_p2 else row['SN P2']

        return sn_add_value, sn_p2_value

    df_maxwire[['SN ADD', 'SN P2']] = df_maxwire.apply(lambda row: update_sn_add_and_sn_p2(row, df_mmsta5_clean), axis=1, result_type='expand')

    df_maxwire['SN ADD'] = df_maxwire['SN ADD'].apply(lambda x: x if isinstance(x, list) else [x] if pd.notna(x) else [])
    df_maxwire['SN P2'] = df_maxwire['SN P2'].apply(lambda x: x if isinstance(x, list) else [x] if pd.notna(x) else [])

    df_maxwire = df_maxwire.explode(['SN ADD', 'SN P2']).reset_index(drop=True)

    df_maxwire['TYPE'] = df_maxwire.apply(
        lambda row: 'Joint' if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '' and (pd.isna(row['TYPE']) or str(row['TYPE']).strip() == '') else row['TYPE'],
        axis=1
    )

    columns_to_check = df_mmsta5.columns[df_mmsta5.columns.get_loc('DSN3'):df_mmsta5.columns.get_loc('Total') + 1]
    columns_to_update = df_maxwire.columns[df_maxwire.columns.get_loc('To Eyelet Protection'):df_maxwire.columns.get_loc('salma') + 1]

    for index, row in df_maxwire.iterrows():
        salma = row['salma']
        matched_row = df_mmsta5_clean[df_mmsta5_clean['DSN2'] == salma]

        if not matched_row.empty:
            for col_to_check, col_to_update in zip(columns_to_check, columns_to_update):
                if matched_row.iloc[0][col_to_check] == 1:
                    df_maxwire.loc[index, col_to_update] = 'X'
                else:
                    df_maxwire.loc[index, col_to_update] = np.nan




    # Étape 1: Nettoyage des données dans df_mmsta pour ignorer les SN1 vides ou contenant "(blanks)"
    df_mmsta_clean = df_mmsta.dropna(subset=['SN1']).copy()  # Supprime les lignes où SN1 est NaN
    df_mmsta_clean = df_mmsta_clean[df_mmsta_clean['SN1'].str.strip() != '(blanks)']  # Supprime les lignes où SN1 est "(blanks)"

    # Étape 2: Fonction pour mapper 'SN ADD' en utilisant 'salma' et 'DSN1'
    def map_sn_add(salma, df):
        match = df.loc[df['DSN1'] == salma, 'SN1']
        return match.values[0] if not match.empty else None

    # Fonction pour mettre à jour 'SN ADD' uniquement si elle est vide
    def update_sn_add_if_empty(row, df_mmsta_clean):
        # Conserver la valeur existante si elle contient des valeurs non vides
        if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '':
            return row['SN ADD']
        # Sinon, essayer de remplir 'SN ADD' avec les correspondances trouvées
        salma = row['salma']
        new_sn_add = map_sn_add(salma, df_mmsta_clean)
        return new_sn_add if new_sn_add else row['SN ADD']

    # Étape 3: Mettre à jour la colonne 'SN ADD' uniquement si elle est vide
    df_maxwire['SN ADD'] = df_maxwire.apply(lambda row: update_sn_add_if_empty(row, df_mmsta_clean), axis=1)

    # Étape 4: Aplatir la liste des SN ADD (si nécessaire)
    df_maxwire['SN ADD'] = df_maxwire['SN ADD'].apply(lambda x: x if isinstance(x, list) else [x] if pd.notna(x) else [])
    df_maxwire = df_maxwire.explode('SN ADD').reset_index(drop=True)

    # Étape 5: Remplissage de la colonne 'TYPE' avec 'FIL SIMPLE' uniquement si elle est vide et que 'SN ADD' est non vide
    df_maxwire['TYPE'] = df_maxwire.apply(
        lambda row: 'FIL SIMPLE' if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '' and (pd.isna(row['TYPE']) or str(row['TYPE']).strip() == '') else row['TYPE'],
        axis=1
    )

    # Étape 6: Vérifier le nombre de lignes avec 'TYPE' égal à 'FIL SIMPLE'
    num_fil_simple = df_maxwire[df_maxwire['TYPE'] == 'FIL SIMPLE'].shape[0]
    num_separation_rows = df_mmsta_clean.shape[0]
    print(f"Initial count of 'FIL SIMPLE': {num_fil_simple}")
    print(f"Number of rows in df_mmsta_clean: {num_separation_rows}")

    # Étape 7: Identifier les lignes de df_mmsta_clean qui ne sont pas encore intégrées dans df_maxwire
    non_integrated_df = df_mmsta_clean[~df_mmsta_clean['SN1'].isin(df_maxwire['SN ADD'])]

    # Préparer les caractéristiques à vérifier dans 'DS Général'
    characteristics = ['From Seal by Terminal', 'From Terminal', 'To Terminal', 'To Seal by Terminal', 'Wire Part Number', 'Final Wire Length']

    # S'assurer que toutes les colonnes pertinentes sont traitées comme des chaînes et gérer les valeurs NaN
    for char in characteristics:
        df_maxwire.loc[:,char] = df_maxwire[char].astype(str).fillna('')

    non_integrated_df.loc[:,'DS Général'] = non_integrated_df['DS Général'].astype(str).fillna('')

    # Fonction pour vérifier si toutes les caractéristiques pertinentes sont des sous-chaînes de la colonne DS Général
    def check_inclusion(row, ds_general):
        items = [row[char].strip() for char in characteristics if row[char].strip() not in ['', 'nan']]
        return all(item in ds_general for item in items)

    # Fonction pour intégrer les valeurs de 'SN ADD' basées sur les caractéristiques, sans écrasement
    def integrate_sn_add(row, non_integrated_df):
        if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '':
            return row['SN ADD']  # Conserver la valeur existante si elle contient des valeurs non vides
        matched_rows = non_integrated_df[non_integrated_df['DS Général'].apply(lambda x: check_inclusion(row, x))]
        if not matched_rows.empty:
            return matched_rows['SN1'].tolist()  # Retourner toutes les correspondances sous forme de liste
        return []  # Retourner une liste vide si aucune correspondance trouvée

    # Appliquer l'intégration pour les lignes non intégrées uniquement à partir de non_integrated_df
    df_maxwire['SN ADD'] = df_maxwire.apply(lambda row: integrate_sn_add(row, non_integrated_df) if row['TYPE'] != 'FIL SIMPLE' else row['SN ADD'], axis=1)

    # Aplatir à nouveau la liste des SN ADD après intégration
    df_maxwire = df_maxwire.explode('SN ADD').reset_index(drop=True)

    # Étape 8: Mettre à jour la colonne 'TYPE' après l'intégration supplémentaire, si 'TYPE' est encore vide
    df_maxwire['TYPE'] = df_maxwire.apply(
        lambda row: 'FIL SIMPLE' if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '' and (pd.isna(row['TYPE']) or str(row['TYPE']).strip() == '') else row['TYPE'],
        axis=1
    )

    # Vérifier le nombre de lignes avec 'TYPE' égal à 'FIL SIMPLE' après intégration
    num_fil_simple_updated = df_maxwire[df_maxwire['TYPE'] == 'FIL SIMPLE'].shape[0]
    print(f"Updated count of 'FIL SIMPLE': {num_fil_simple_updated}")
    print(f"Number of rows in df_mmsta_clean: {num_separation_rows}")

    # Remplacer les valeurs de 'SN ADD' et 'TYPE' par des chaînes vides lorsque la ligne contient "double"
    print("Replacing 'SN ADD' and 'TYPE' where 'double' is found")
    mask = df_maxwire.apply(lambda row: row.astype(str).str.contains('double', case=False).any(), axis=1)
    df_maxwire.loc[mask, ['SN ADD', 'TYPE']] = ''

  


    # Étape 1: Nettoyage des données dans df_mmsta1 pour ignorer les SN2 vides ou contenant "(blanks)"
    df_mmsta1_clean = df_mmsta1.dropna(subset=['SN2']).copy()  # Supprime les lignes où SN2 est NaN
    df_mmsta1_clean = df_mmsta1_clean[df_mmsta1_clean['SN2'].str.strip() != '(blanks)']  # Supprime les lignes où SN2 est "(blanks)"

    # Étape 2: Remplissage initial de la colonne 'SN ADD' en utilisant une correspondance basée sur 'salma' et 'DSN2'
    def map_sn_add(salma, df):
        matches = df[df['DSN2'] == salma]['SN2']
        return matches.iloc[0] if not matches.empty else ''  # Garder seulement la première correspondance

    # Conserver les valeurs existantes dans 'SN ADD' si elles sont déjà non vides
    def update_sn_add_if_empty(row, df_mmsta1_clean):
        if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '':
            return row['SN ADD']  # Conserver la valeur existante si non vide
        # Sinon, essayer de remplir 'SN ADD' avec la correspondance trouvée
        salma = row['salma']
        if salma in df_mmsta1_clean['DSN2'].values:
            return map_sn_add(salma, df_mmsta1_clean)
        return row.get('SN ADD', '')  # Retourner la valeur existante ou une chaîne vide si aucune correspondance trouvée

    # Étape 3: Remplir la colonne 'SN ADD'
    df_maxwire['SN ADD'] = df_maxwire.apply(lambda row: update_sn_add_if_empty(row, df_mmsta1_clean), axis=1)

    # Étape 4: Mise à jour de la colonne 'TYPE'
    df_maxwire['TYPE'] = df_maxwire.apply(
        lambda row: 'Double' if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '' and (pd.isna(row['TYPE']) or str(row['TYPE']).strip() == '') else row['TYPE'],
        axis=1
    )

    # Étape 5: Conserver les valeurs de 'SN ADD' existantes pour les lignes où 'SN ADD' est déjà non vide
    def integrate_sn_add(row, non_integrated_df1):
        if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '':
            return row['SN ADD']  # Conserver la valeur existante si non vide
        matched_rows = non_integrated_df1[non_integrated_df1['DS Général'].apply(lambda x: check_inclusion(row, x))]
        if not matched_rows.empty:
            return matched_rows['SN2'].iloc[0]
        return ''  # Retourner une chaîne vide si aucune correspondance trouvée

    # Identifier les lignes non intégrées
    non_integrated_df1 = df_mmsta1_clean[~df_mmsta1_clean['SN2'].isin(df_maxwire['SN ADD'])]

    # Préparation des caractéristiques (conversion en chaînes de caractères)
    characteristics = ['From Seal by Terminal', 'From Terminal', 'To Terminal', 'To Seal by Terminal', 'Wire Part Number', 'Final Wire Length']

    for char in characteristics:
        df_maxwire.loc[:,char] = df_maxwire[char].astype(str).fillna('')

    non_integrated_df1.loc[:,'DS Général'] = non_integrated_df1['DS Général'].astype(str).fillna('')

    # Vérification de l'inclusion des caractéristiques
    def check_inclusion(row, ds_general):
        items = [row[char].strip() for char in characteristics if row[char].strip() not in ['', 'nan']]
        return all(item in ds_general for item in items)

    # Mettre à jour les valeurs de 'SN ADD' uniquement si elles sont vides
    df_maxwire['SN ADD'] = df_maxwire.apply(lambda row: integrate_sn_add(row, non_integrated_df1), axis=1)

    # Étape 6: Mettre à jour 'SN P2' là où 'SN ADD' est égale à 'SN2' de df_mmsta1
    for i, row in df_maxwire[df_maxwire['SN ADD'].notna()].iterrows():
        sn_add = row['SN ADD']
        sn_1_value = df_mmsta1.loc[df_mmsta1['SN2'] == sn_add, 'SN1']
        if not sn_1_value.empty:
            df_maxwire.at[i, 'SN P2'] = sn_1_value.values[0]




    # Étape 1: Nettoyage des données dans df_mmsta2 pour ignorer les SN2 vides ou contenant "(blanks)"
    df_mmsta2_clean = df_mmsta2.dropna(subset=['SN2']).copy()  # Supprime les lignes où SN2 est NaN
    df_mmsta2_clean = df_mmsta2_clean[df_mmsta2_clean['SN2'].str.strip() != '(blanks)']  # Supprime les lignes où SN2 est "(blanks)"

    # Étape 2: Remplissage de la colonne 'SN ADD' uniquement si elle est vide
    def map_sn_add(salma, df):
        matches = df[df['DSN2'] == salma]['SN2']
        return matches.iloc[0] if not matches.empty else ''

    # Conserver les valeurs existantes dans 'SN ADD' si elles ne sont pas vides
    def update_sn_add_if_empty(row, df_mmsta2_clean):
        if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '':
            return row['SN ADD']  # Conserver la valeur existante si non vide
        # Sinon, essayer de remplir 'SN ADD' avec la correspondance trouvée
        salma = row['salma']
        if salma in df_mmsta2_clean['DSN2'].values:
            return map_sn_add(salma, df_mmsta2_clean)
        return row.get('SN ADD', '')

    # Étape 3: Remplir la colonne 'SN ADD'
    df_maxwire['SN ADD'] = df_maxwire.apply(lambda row: update_sn_add_if_empty(row, df_mmsta2_clean), axis=1)

    # Étape 4: Mise à jour de la colonne 'TYPE'
    df_maxwire['TYPE'] = df_maxwire.apply(
        lambda row: 'Twist' if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '' and (pd.isna(row['TYPE']) or str(row['TYPE']).strip() == '') else row['TYPE'],
        axis=1
    )

    # Étape 5: Vérification des 'Twist'
    num_twist = df_maxwire[df_maxwire['TYPE'] == 'Twist'].shape[0]
    print(f"Count of 'Twist': {num_twist}")

    # Étape 6: Identification des lignes 'Twist' et non intégrées
    twist_df = df_maxwire[df_maxwire['TYPE'] == 'Twist'].copy()
    non_integrated_df2 = df_mmsta2_clean[~df_mmsta2_clean['SN2'].isin(twist_df['SN ADD'])]

    # Étape 7: Préparation des caractéristiques (conversion en chaînes de caractères)
    characteristics = ['From Seal by Terminal', 'From Terminal', 'To Terminal', 'To Seal by Terminal', 'Wire Part Number', 'Final Wire Length']

    for char in characteristics:
        df_maxwire.loc[:,char] = df_maxwire[char].astype(str).fillna('')

    non_integrated_df2.loc[:,'DS Général'] = non_integrated_df2['DS Général'].astype(str).fillna('')

    # Vérification de l'inclusion des caractéristiques
    def check_inclusion(row, ds_general):
        items = [row[char].strip() for char in characteristics if row[char].strip() not in ['', 'nan']]
        return all(item in ds_general for item in items)

    # Étape 8: Intégration de 'SN ADD' pour les lignes non intégrées
    def integrate_sn_add(row, non_integrated_df2):
        if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '':
            return row['SN ADD']  # Conserver la valeur existante si non vide
        matched_rows = non_integrated_df2[non_integrated_df2['DS Général'].apply(lambda x: check_inclusion(row, x))]
        if not matched_rows.empty:
            return matched_rows['SN2'].iloc[0]
        return ''  # Retourner une chaîne vide si aucune correspondance trouvée

    # Mettre à jour les valeurs de 'SN ADD' uniquement si elles sont vides
    df_maxwire['SN ADD'] = df_maxwire.apply(lambda row: integrate_sn_add(row, non_integrated_df2), axis=1)

    # Étape 9: Mettre à jour 'SN P2' là où 'TYPE' est 'Twist' et 'SN ADD' est égale à 'SN2' de df_mmsta2
    for i, row in df_maxwire[df_maxwire['TYPE'] == 'Twist'].iterrows():
        sn_add = row['SN ADD']
        sn_1_value = df_mmsta2.loc[df_mmsta2['SN2'] == sn_add, 'SN1']
        if not sn_1_value.empty:
            df_maxwire.at[i, 'SN P2'] = sn_1_value.values[0]



    # Étape 1: Nettoyage des données dans df_mmsta3 pour ignorer les SN2 vides ou contenant "(blanks)"
    df_mmsta3_clean = df_mmsta3.dropna(subset=['SN2']).copy()  # Supprime les lignes où SN2 est NaN
    df_mmsta3_clean = df_mmsta3_clean[df_mmsta3_clean['SN2'].str.strip() != '(blanks)']  # Supprime les lignes où SN2 est "(blanks)"

    # Étape 2: Remplissage de la colonne 'SN ADD' uniquement si elle est vide
    def map_sn_add(salma, df):
        matches = df[df['DSN2'] == salma]['SN2']
        return matches.iloc[0] if not matches.empty else ''

    # Conserver les valeurs existantes dans 'SN ADD' si elles ne sont pas vides
    def update_sn_add_if_empty(row, df_mmsta3_clean):
        if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '':
            return row['SN ADD']  # Conserver la valeur existante si non vide
        # Sinon, essayer de remplir 'SN ADD' avec la correspondance trouvée
        salma = row['salma']
        if salma in df_mmsta3_clean['DSN2'].values:
            return map_sn_add(salma, df_mmsta3_clean)
        return row.get('SN ADD', '')

    # Étape 3: Remplir la colonne 'SN ADD'
    df_maxwire['SN ADD'] = df_maxwire.apply(lambda row: update_sn_add_if_empty(row, df_mmsta3_clean), axis=1)

    # Étape 4: Mise à jour de la colonne 'TYPE'
    df_maxwire['TYPE'] = df_maxwire.apply(
        lambda row: 'SQUIB' if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '' and (pd.isna(row['TYPE']) or str(row['TYPE']).strip() == '') else row['TYPE'],
        axis=1
    )

    # Étape 5: Vérification des 'SQUIB'
    num_squib = df_maxwire[df_maxwire['TYPE'] == 'SQUIB'].shape[0]
    print(f"Count of 'SQUIB': {num_squib}")

    # Étape 6: Identification des lignes 'SQUIB' et non intégrées
    squib_df = df_maxwire[df_maxwire['TYPE'] == 'SQUIB'].copy()
    non_integrated_df3 = df_mmsta3_clean[~df_mmsta3_clean['SN2'].isin(squib_df['SN ADD'])]

    # Étape 7: Préparation des caractéristiques (conversion en chaînes de caractères)
    characteristics = ['From Seal by Terminal', 'From Terminal', 'To Terminal', 'To Seal by Terminal', 'Wire Part Number', 'Final Wire Length']

    for char in characteristics:
        df_maxwire.loc[:,char] = df_maxwire[char].astype(str).fillna('')

    non_integrated_df3.loc[:,'DS Général'] = non_integrated_df3['DS Général'].astype(str).fillna('')

    # Vérification de l'inclusion des caractéristiques
    def check_inclusion(row, ds_general):
        items = [row[char].strip() for char in characteristics if row[char].strip() not in ['', 'nan']]
        return all(item in ds_general for item in items)

    # Étape 8: Intégration de 'SN ADD' pour les lignes non intégrées
    def integrate_sn_add(row, non_integrated_df3):
        if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '':
            return row['SN ADD']  # Conserver la valeur existante si non vide
        matched_rows = non_integrated_df3[non_integrated_df3['DS Général'].apply(lambda x: check_inclusion(row, x))]
        if not matched_rows.empty:
            return matched_rows['SN2'].iloc[0]
        return ''  # Retourner une chaîne vide si aucune correspondance trouvée

    # Mettre à jour les valeurs de 'SN ADD' uniquement si elles sont vides
    df_maxwire['SN ADD'] = df_maxwire.apply(lambda row: integrate_sn_add(row, non_integrated_df3), axis=1)

    # Étape 9: Mettre à jour 'SN P2' là où 'TYPE' est 'SQUIB' et 'SN ADD' est égale à 'SN2' de df_mmsta3
    for i, row in df_maxwire[df_maxwire['TYPE'] == 'SQUIB'].iterrows():
        sn_add = row['SN ADD']
        sn_1_value = df_mmsta3.loc[df_mmsta3['SN2'] == sn_add, 'SN1']
        if not sn_1_value.empty:
            df_maxwire.at[i, 'SN P2'] = sn_1_value.values[0]


    # Étape 1: Nettoyage des données dans df_mmsta4 pour ignorer les SN2 vides ou contenant "(blanks)"
    df_mmsta4_clean = df_mmsta4.dropna(subset=['SN2']).copy()  # Supprime les lignes où SN2 est NaN
    df_mmsta4_clean = df_mmsta4_clean[df_mmsta4_clean['SN2'].str.strip() != '(blanks)']  # Supprime les lignes où SN2 est "(blanks)"

    # Étape 2: Remplissage de la colonne 'SN ADD' uniquement si elle est vide
    def map_sn_add(salma, df):
        matches = df[df['DSN2'] == salma]['SN2']
        return matches.iloc[0] if not matches.empty else ''

    # Conserver les valeurs existantes dans 'SN ADD' si elles ne sont pas vides
    def update_sn_add_if_empty(row, df_mmsta4_clean):
        if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '':
            return row['SN ADD']  # Conserver la valeur existante si non vide
        # Sinon, essayer de remplir 'SN ADD' avec la correspondance trouvée
        salma = row['salma']
        if salma in df_mmsta4_clean['DSN2'].values:
            return map_sn_add(salma, df_mmsta4_clean)
        return row.get('SN ADD', '')

    # Étape 3: Remplir la colonne 'SN ADD'
    df_maxwire['SN ADD'] = df_maxwire.apply(lambda row: update_sn_add_if_empty(row, df_mmsta4_clean), axis=1)

    # Étape 4: Mise à jour de la colonne 'TYPE'
    df_maxwire['TYPE'] = df_maxwire.apply(
        lambda row: 'GW' if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '' and (pd.isna(row['TYPE']) or str(row['TYPE']).strip() == '') else row['TYPE'],
        axis=1
    )

    # Étape 5: Vérification des 'GW'
    num_gw = df_maxwire[df_maxwire['TYPE'] == 'GW'].shape[0]
    print(f"Count of 'GW': {num_gw}")

    # Étape 6: Identification des lignes 'GW' et non intégrées
    gw_df = df_maxwire[df_maxwire['TYPE'] == 'GW'].copy()
    non_integrated_df4 = df_mmsta4_clean[~df_mmsta4_clean['SN2'].isin(gw_df['SN ADD'])]

    # Étape 7: Préparation des caractéristiques (conversion en chaînes de caractères)
    characteristics = ['From Seal by Terminal', 'From Terminal', 'To Terminal', 'To Seal by Terminal', 'Wire Part Number', 'Final Wire Length']

    for char in characteristics:
        df_maxwire.loc[:,char] = df_maxwire[char].astype(str).fillna('')

    non_integrated_df4.loc[:,'DS Général'] = non_integrated_df4['DS Général'].astype(str).fillna('')

    # Vérification de l'inclusion des caractéristiques
    def check_inclusion(row, ds_general):
        items = [row[char].strip() for char in characteristics if row[char].strip() not in ['', 'nan']]
        return all(item in ds_general for item in items)

    # Étape 8: Intégration de 'SN ADD' pour les lignes non intégrées
    def integrate_sn_add(row, non_integrated_df4):
        if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '':
            return row['SN ADD']  # Conserver la valeur existante si non vide
        matched_rows = non_integrated_df4[non_integrated_df4['DS Général'].apply(lambda x: check_inclusion(row, x))]
        if not matched_rows.empty:
            return matched_rows['SN2'].iloc[0]
        return ''  # Retourner une chaîne vide si aucune correspondance trouvée

    # Mettre à jour les valeurs de 'SN ADD' uniquement si elles sont vides
    df_maxwire['SN ADD'] = df_maxwire.apply(lambda row: integrate_sn_add(row, non_integrated_df4), axis=1)

    # Étape 9: Mettre à jour 'SN P2' là où 'TYPE' est 'GW' et 'SN ADD' est égale à 'SN2' de df_mmsta4
    for i, row in df_maxwire[df_maxwire['TYPE'] == 'GW'].iterrows():
        sn_add = row['SN ADD']
        sn_1_value = df_mmsta4.loc[df_mmsta4['SN2'] == sn_add, 'SN1']
        if not sn_1_value.empty:
            df_maxwire.at[i, 'SN P2'] = sn_1_value.values[0]




   












    df_mmsta6_clean = df_mmsta6.dropna(subset=['SN3']).copy()
    df_mmsta6_clean = df_mmsta6_clean[df_mmsta6_clean['SN3'].str.strip() != '(blanks)']

    def map_sn_add_sn_p2_and_supergroup(salma, df):
        matched_row = df[df['DSN3'] == salma]
        if not matched_row.empty:
            sn_add = matched_row['SN3'].tolist()
            sn_p2 = matched_row['SN2'].tolist()
            supergroup = matched_row['SN1'].tolist()
            return sn_add, sn_p2, supergroup
        return None, None, None

    def update_sn_add_sn_p2_and_supergroup(row, df_mmsta6_clean):
        salma = row['salma']
        new_sn_add, new_sn_p2, new_supergroup = map_sn_add_sn_p2_and_supergroup(salma, df_mmsta6_clean)

        if isinstance(row['SN ADD'], list) and any(str(val).strip() for val in row['SN ADD']):
            sn_add_value = row['SN ADD']
        else:
            sn_add_value = new_sn_add if new_sn_add else row['SN ADD']

        if isinstance(row['SN P2'], list) and any(str(val).strip() for val in row['SN P2']):
            sn_p2_value = row['SN P2']
        else:
            sn_p2_value = new_sn_p2 if new_sn_p2 else row['SN P2']

        if isinstance(row['Super group'], list) and any(str(val).strip() for val in row['Super group']):
            supergroup_value = row['Super group']
        else:
            supergroup_value = new_supergroup if new_supergroup else row['Super group']

        return sn_add_value, sn_p2_value, supergroup_value

    df_maxwire[['SN ADD', 'SN P2', 'Super group']] = df_maxwire.apply(lambda row: update_sn_add_sn_p2_and_supergroup(row, df_mmsta6_clean), axis=1, result_type='expand')

    df_maxwire['SN ADD'] = df_maxwire['SN ADD'].apply(lambda x: x if isinstance(x, list) else [x] if pd.notna(x) else [])
    df_maxwire['SN P2'] = df_maxwire['SN P2'].apply(lambda x: x if isinstance(x, list) else [x] if pd.notna(x) else [])
    df_maxwire['Super group'] = df_maxwire['Super group'].apply(lambda x: x if isinstance(x, list) else [x] if pd.notna(x) else [])

    df_maxwire = df_maxwire.explode(['SN ADD', 'SN P2', 'Super group']).reset_index(drop=True)

    df_maxwire['TYPE'] = df_maxwire.apply(
        lambda row: 'SG' if pd.notna(row['SN ADD']) and str(row['SN ADD']).strip() != '' and (pd.isna(row['TYPE']) or str(row['TYPE']).strip() == '') else row['TYPE'],
        axis=1
    )

    columns_to_check = df_mmsta6.columns[df_mmsta6.columns.get_loc('DSN3'):df_mmsta6.columns.get_loc('Total') + 1]
    columns_to_update = df_maxwire.columns[df_maxwire.columns.get_loc('To Eyelet Protection'):df_maxwire.columns.get_loc('salma') + 1]

    for index, row in df_maxwire.iterrows():
        salma = row['salma']
        matched_row = df_mmsta6_clean[df_mmsta6_clean['DSN3'] == salma]

        if not matched_row.empty:
            for col_to_check, col_to_update in zip(columns_to_check, columns_to_update):
                if matched_row.iloc[0][col_to_check] == 1:
                    df_maxwire.loc[index, col_to_update] = 'X'
                else:
                    df_maxwire.loc[index, col_to_update] = np.nan

    df_maxwire = df_maxwire[~((df_maxwire['TYPE'] == 'SG') & df_maxwire.duplicated(subset=['SN ADD', 'SN P2', 'Super group'], keep='first'))]

    output_path = 'liste_circuit_integre.xlsx'
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_maxwire.to_excel(writer, sheet_name='Integrated Data', index=False)
        types = ['FIL SIMPLE', 'Double', 'Twist', 'SQUIB', 'GW', 'Joint', 'SG']
        type_counts_maxwire = {t: len(df_maxwire[df_maxwire['TYPE'] == t]) for t in types}

        type_counts_mmsta = {
            'FIL SIMPLE': df_mmsta['DS Général'].dropna().count(),
            'Double': df_mmsta1['DS Général'].dropna().count(),
            'Twist': df_mmsta2['DS Général'].dropna().count(),
            'SQUIB': df_mmsta3['DS Général'].dropna().count(),
            'GW': df_mmsta4['DS Général'].dropna().count(),
            'Joint': df_mmsta5['DS Général'].dropna().count(),
            'SG': df_mmsta6['DS Général'].dropna().count()
        }

        comparison_df = pd.DataFrame({
            'Type': types,
            'Nombre de lignes dans df_mmsta': [type_counts_mmsta[t] for t in types],
            'Nombre de lignes dans df_maxwire': [type_counts_maxwire[t] for t in types]
        })

        comparison_df.to_excel(writer, sheet_name='Comparison', index=False)

    return output_path

# Streamlit app setup
st.set_page_config(page_title="YMM-LASA for Automated Integration", page_icon=":calendar:", layout="wide")

with st.sidebar:
    menu = st.selectbox("Select an option", ["Home", "Part 1: Separation", "Part 2: Integration", "Help"])

    st.markdown("""
        <div style='position: fixed; bottom: 0; left: 0; padding: 10px;'>
            <p>Developed by <strong>EL MALIANI LATIFA</strong> and <strong>EL BINANI SALMA</strong></p>
        </div>
        """, unsafe_allow_html=True)

def add_bg_from_local(image_file):
    with open(image_file, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read())
    st.markdown(
        f"""
        <style>
        .stApp {{
        background-image: url(data:image/{"png"};base64,{encoded_string.decode()});
        background-size: cover
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

if menu == "Home":
    add_bg_from_local('image.png')
    st.markdown("""
        <div style='text-align: center; padding: 20px;'>
            <h1 style="font-size: 2.5em; color: #007bff; animation: fadeIn 2s;">YMM-LASA Automated Application</h1>
            <p style="font-size: 1.2em; color: #333; margin-top: 20px;">
                Welcome to the world of <strong>"YMM-LASA Automated Application"</strong>, a cutting-edge technological solution designed to transform circuit management within the Production Preparation (PP) department at Yazaki Meknès. This application is the result of a collaboration between Salma El Binani and Latifa El Maliani, two engineers specializing in artificial intelligence and data science.
            </p>
            <p style="font-size: 1.2em; color: #333; margin-top: 20px;">
                <strong>Intelligent Automation for Optimized Production</strong><br>
                Through the advanced use of the MMSTA file, generated by SAP, "YMM-LASA Automated Application" automates the separation of circuit types. This process, which was once manual and time-consuming, is now streamlined, allowing operators to work more efficiently and accurately.
            </p>
            <p style="font-size: 1.2em; color: #333; margin-top: 20px;">
                <strong>Data Integration for Precise Management</strong><br>
                One of the key strengths of this application lies in its ability to seamlessly integrate data from the MMSTA file directly into the LIST CIRCUIT file, produced by the PTC (Product and Tooling Configuration) tool. This integration consolidates all circuit-related information into a single location, ensuring more precise and comprehensive management of production elements.
            </p>
            <p style="font-size: 1.2em; color: #333; margin-top: 20px;">
                <strong>A Tangible Impact on Productivity</strong><br>
                The impact of "YMM-LASA Automated Application" on the PP department is immediate and significant.
            </p>
        </div>
        <style>
        @keyframes fadeIn {{
            from {{ opacity: 0; }}
            to {{ opacity: 1; }}
        }}
        </style>
        """, unsafe_allow_html=True)

current_date = get_current_time()
expiration_date = datetime(2024, 9, 10, 14, 0, tzinfo=pytz.UTC)

if current_date > expiration_date:
    st.error("Thank you for your visit.")
else:
    st.success("Welcome to the application!")

    uploads_dir = "uploads"
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)

    if menu == "Part 1: Separation":
        add_bg_from_local('Template.jpg')
        st.header("Part 1: Separation")

        uploaded_mmsta_file = st.file_uploader("Upload your MMSTA file", type=["csv", "xlsx"])

        if st.button("Separate"):
            if uploaded_mmsta_file:
                mmsta_filepath = os.path.join(uploads_dir, uploaded_mmsta_file.name)
                with open(mmsta_filepath, "wb") as f:
                    f.write(uploaded_mmsta_file.getbuffer())

                output_file = separate_files(mmsta_filepath)
                if output_file:
                    with open(output_file, 'rb') as generated_file:
                        st.download_button(
                            label="Download the separated Excel file",
                            data=generated_file,
                            file_name=output_file,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    st.success("Separation completed successfully!")
                else:
                    st.error("Error generating the Excel file.")
            else:
                st.error("Please upload the MMSTA file.")

    elif menu == "Part 2: Integration":
        add_bg_from_local('Template.jpg')
        st.header("Part 2: Integration")

        uploaded_mmsta_separe_file = st.file_uploader("Upload the separated MMSTA file (MMSTA_separe.xlsx)", type=["xlsx"])
        uploaded_circuit_file = st.file_uploader("Upload your Circuit List file", type=["csv", "xlsx"])

        if st.button("Integrate"):
            if uploaded_mmsta_separe_file and uploaded_circuit_file:
                mmsta_separe_filepath = os.path.join(uploads_dir, uploaded_mmsta_separe_file.name)
                with open(mmsta_separe_filepath, "wb") as f:
                    f.write(uploaded_mmsta_separe_file.getbuffer())
                
                circuit_filepath = os.path.join(uploads_dir, uploaded_circuit_file.name)
                with open(circuit_filepath, "wb") as f:
                    f.write(uploaded_circuit_file.getbuffer())

                output_file = integrate_files(mmsta_separe_filepath, circuit_filepath)
                if output_file:
                    with open(output_file, 'rb') as integration_file:
                        st.download_button(
                            label="Download the integrated Excel file",
                            data=integration_file,
                            file_name=output_file,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    st.success("Integration completed successfully!")
                else:
                    st.error("Error generating the Excel file.")
            else:
                st.error("Please upload both the separated MMSTA file and the Circuit List.")

    elif menu == "Help":
        add_bg_from_local('Template.jpg')
        st.header("Help")
        st.write("""
            *Part 1: Separation* - Use this section to upload an MMSTA file and separate it into different components.
            
            *Part 2: Integration* - Upload the separated MMSTA file and a Circuit List file to integrate them into a single output.
            
            *Help* - This section provides information about how to use the application.
        """)
